# 不能处理xls
from openpyxl import load_workbook
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import urllib.request
import ssl
import re
from collections import OrderedDict  # 有序字典
from pyexcel_xls import save_data  # 读取数据
from openpyxl.reader.excel import load_workbook


def getHtmlBytes(url):
    headers = {  # 模拟请求头
        "Accept": "application/jaon,text/javascript, */*;q=0.01",
        "X-Requested-With": "XMLHttpRequest",  # 请求对象类型
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36",
        "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8"  # 链接类型

    }
    req = urllib.request.Request(url, headers=headers)
    context = ssl._create_unverified_context()
    response = urllib.request.urlopen(req, context=context)
    return response.read()


# 一般分好几页或者好几个项目，可以让返回会dict
# {“第一个标题（即key）”：[[ （这里是第一行第一列）, , ,](这里面存放第一行数据),[, , ,],[, , ,]]}
def readXlsxFile(path, path2):
    # 打开文件
    file = load_workbook(filename=path)
    # 所有表格的名称,list类型返回
    sheets = file.sheetnames  # gget_sheet_names已经被弃用这里用sheetnames代替
    # 拿出一个表格
    '''
    sheet = sheets[0]
    print(type(sheet))sheet是一个str类型,因为list存的是str'''
    # get_by_sheet_name已经弃用用file[file.sheetnames[0]]代替或者sheet=["sheet1"]
    sheet = file[file.sheetnames[0]]
    # 最大行数
    # print(sheet.max_row)
    # 最大列数
    # print(sheet.max_column)
    # 表名
    # print(sheet.tittle)
    # 读取数据
    for lineNum in range(1, sheet.max_row + 1):  # 注意这里从1开始
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            value = sheet.cell(row=lineNum, column=columnNum).value
            # if value != None:#none用来判断是否为空格或者没有看数据可以不用，写数据一定要用
            lineList.append(value)
        # print(lineList[1])lineList[1]是后缀
        # print(lineList)
        url = "https://www.kegg.jp/dbget-bin/www_bfind_sub?mode=bfind&max_hit=1000&dbkey=kegg&keywords=" + str(
            lineList[1])
        # print(url)
        htmlBytes = getHtmlBytes(url)
        # writeFileBytes(htmlBytes,r"C:\Users\我爱陆地\PycharmProjects\爬虫简介与json\爬虫\爬QQ号\qqFile1.html")
        # writerFileStr(htmlBytes,r"C:\Users\我爱陆地\PycharmProjects\爬虫简介与json\爬虫\爬QQ号\qqFile2.txt")
        # 到这里说明趴下了没问题
        htmlStr = str(htmlBytes)

        pat = r'K.{5}</a><br><div style="margin-left:2em">.*</div></div><br><div style="color'
        re_qq = re.compile(pat)
        qqsList = re_qq.findall(htmlStr)
        # print(qqsList)
        # print(type(qqsList))
        # print(type(lineList[1]))
        # print(lineList[1]+str(qqsList))
        list2 = qqsList[0].split('">')
        list3 = list2[1].split('<', 1)
        # print(list3[0])
        lineList.append(list3[0])
        print(lineList)
        # lineList=dict(lineList)
        # print(type(list3[0]))


path = r"D:\My Documents\Tencent Files\1846851491\FileRecv\KEGG.xlsx"
path2 = r"D:\My Documents\Tencent Files\1846851491\FileRecv\KEEG!.XLSX"
readXlsxFile(path, path2)